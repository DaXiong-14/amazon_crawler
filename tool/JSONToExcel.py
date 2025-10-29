import logging
import time
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import  PatternFill, Alignment
import json
import os
import requests
from io import BytesIO
from Baidu_Text_transAPI import BaiduTranslation
from openpyxl.styles import colors
from openpyxl.styles import Font


logger = logging.getLogger(__name__)


class AmazonExcelExporter:
    def __init__(self, filename="amazon_products.xlsx"):
        """
        初始化Excel导出器
        :param filename: 输出Excel文件名
        """
        self.filename = filename
        self.wb = None
        self.ws = None
        self.headers = [
            "ASIN", "品牌", "图片", "售价", "评分", "BSR/NR排名", "子类目链接",
            "月销量", "月销额", "上架时间", "变体数量", "重量", "包装尺寸",
            "卖家", "卖家地址", "补充信息", "同款信息", "同款图一", "同款图二",
            "同款图三", "预计配送费", "五点描述", "图1","图2", "图3", "采购价(元)",
            "采购商(1688)", "供应商1688链接"
        ]
        self.column_widths = {
            'A': 15, 'B': 15, 'C': 15, 'D': 10, 'E': 8, 'F': 15, 'G': 15,
            'H': 10, 'I': 12, 'J': 20, 'K': 10, 'L': 12, 'M': 15, "N": 15,
            "O": 50, "P": 30, 'Q': 30, 'R': 15, 'S': 15, 'T': 15, 'U': 12,
            'V': 30, 'W': 15, 'X': 15, 'Y': 15, 'Z': 12, 'AA': 20, 'AB': 30
        }
        # 图片大小配置
        self.image_config = {
            'main_image': {'width': 80, 'height': 80},
            'variant_image': {'width': 60, 'height': 60}
        }
        # 行高配置
        self.row_heights = {
            'data_row': 80  # 数据行高度
        }
        # get_image 重试次数
        self.i_img = 0


    def download_image(self, url):
        """下载图片并返回BytesIO对象"""
        try:
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
            }
            response = requests.get(url, headers=headers, timeout=10, verify=False)
            response.raise_for_status()
            self.i_img = 0
            return BytesIO(response.content)
        except Exception as e:
            logger.error(f"下载图片失败 {url}: {e}, 正在重试")
            self.i_img += 1
            if self.i_img >= 5:
                # 返回一个简单的占位符
                return BytesIO(
                    b'iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR42mP8/5+hHgAHggJ/PchI7wAAAABJRU5ErkJggg==')
            return self.download_image(url)

    def create_worksheet(self, sheet_name="产品数据"):
        """
        创建工作表
        :param sheet_name: 工作表名称
        """
        if os.path.exists(self.filename):
            self.wb = load_workbook(self.filename)
            if sheet_name in self.wb.sheetnames:
                self.ws = self.wb[sheet_name]
            else:
                self.ws = self.wb.create_sheet(title=sheet_name)
        else:
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.title = sheet_name

        # 写入表头（如果工作表是新建的）
        if self.ws.max_row == 1:
            for col, header in enumerate(self.headers, 1):
                cell = self.ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # 设置列宽
        for col, width in self.column_widths.items():
            self.ws.column_dimensions[col].width = width

    def add_product_data(self, item):
        """
        添加产品数据到工作表
        :param item: 产品数据字典
        """
        if not self.wb or not self.ws:
            raise Exception("请先调用create_worksheet方法创建工作表")

        row_idx = self.ws.max_row + 1

        # todo 设置行高
        self.ws.row_dimensions[row_idx].height = self.row_heights['data_row']

        # todo 创建 asin 文本超链接
        asin = item.get("asin", "")
        if asin:
            amazon_url = f"{item.get('itemWEB')}/dp/{asin}?th=1"
            cell = self.ws.cell(row=row_idx, column=1, value=asin)
            cell.hyperlink = amazon_url
            cell.font = Font(color=colors.BLUE, underline='single')

        # todo 品牌
        self.ws.cell(row=row_idx, column=2, value=item.get("brand", ""))

        # todo 主图处理
        image_url = item.get("imageUrl")
        if not image_url:
            image_url = item.get("image")
        if image_url:
            try:
                img_data = self.download_image(image_url)
                img = Image(img_data)
                img.width = self.image_config['main_image']['width']
                img.height = self.image_config['main_image']['height']
                # 修复：直接设置anchor为单元格地址字符串
                img.anchor = f'C{row_idx}'
                self.ws.add_image(img)
            except Exception as e:
                logger.error(f"插入主图失败: {e}")

        # todo 价格
        currency = item.get('currency', '')
        price = item.get("current_price", "")
        if price:
            price = f'现价: {price}\n优惠比: {item.get('discount_percentage', '')}'
        else:
            price = f'{currency}{item.get("price", "")}'
        self.ws.cell(row=row_idx, column=4, value=price)

        # todo 评分
        rating = item.get("rating", "")
        reviewCount = item.get("reviewCount", "")
        if not reviewCount:
            reviewCount = item.get("reviews", "")
        rating_txt = f"{rating} 星\n{reviewCount} 评论"
        self.ws.cell(row=row_idx, column=5, value=rating_txt)

        # todo BSR排名
        bsrText= ''
        bsrList = item.get("bsrList", [])
        if bsrList:
            for bsr in bsrList:
                bsrText += f"#{bsr.get('rank')} {bsr.get('label')}\n"
        self.ws.cell(row=row_idx, column=6, value=bsrText)

        # todo 子类目
        subcategorieText = ''
        subcategories = item.get('subcategories', [])
        if subcategories:
            for subcategory in subcategories:
                subcategorieText += f"{subcategory.get('rank')}: {subcategory.get('label')}\n"
        self.ws.cell(row=row_idx, column=7, value=subcategorieText)

        # todo 销量和销售额
        month_units = item.get("units", "")
        if not month_units:
            item.get('totalUnits')
        self.ws.cell(row=row_idx, column=8, value=str(month_units))

        amount = item.get("amount", "")
        if not amount:
            amount = item.get("totalAmount", "")
        if not amount or float(amount) <= 0:
            amount = 0
        self.ws.cell(row=row_idx, column=9, value=str(amount))

        # todo 上架时间
        itemTime = None
        try:
            updated_time = item.get('available')
            if not updated_time:
                updated_time = item.get('availableDate')
            if updated_time:
                shelves_time = datetime.fromtimestamp(int(updated_time) / 1000)
                formatted_date = shelves_time.strftime('%Y-%m-%d')
                days_on_shelves = (datetime.now() - shelves_time).days
                itemTime = f"{formatted_date} ({days_on_shelves} 天)"
        except Exception as e:
            logger.error(f"时间计算错误: {e}")
        self.ws.cell(row=row_idx, column=10, value=itemTime)

        # todo 变体数量
        self.ws.cell(row=row_idx, column=11, value=item.get("variations", ""))

        # todo 重量和尺寸 (需要换算)
        weight = None
        try:
            # 先检查weight是否存在且不为None
            if item.get("weight") and "Pounds" in item.get("weight"):
                weight_t = item.get("weight").replace(" ", "").replace("Pounds", "")
                weight_f = round(float(weight_t) * 453.59237, 2)
                weight = f"{str(weight_f)} g"
            elif item.get("weight") and "ounces" in item.get("weight"):
                weight_t = item.get("weight").replace(" ", "").replace("ounces", "")
                weight_f = round(float(weight_t) * 28.3495, 2)
                weight = f"{str(weight_f)} g"
            elif item.get("weight"):
                weight = item.get("weight")
            else:
                weight = None  # 如果weight为None，保持为None
        except Exception as e:
            logger.error(f"重量换算错误: {e}")
            weight = None

        # 修复尺寸换算逻辑
        dimension = None
        try:
            # 先检查dimension是否存在且不为None
            if item.get("dimension") and "inches" in item.get("dimension"):
                dimension_t = item.get("dimension").replace(" ", "").replace("inches", "")
                ds = dimension_t.split("x")
                dimension = ""
                for s in ds:
                    dimension += str(round(float(s) * 2.54, 2))
                    dimension += " x "
                dimension = dimension[:-3] + " cm"  # 去掉最后的 " x "
            elif item.get("dimension"):
                dimension = item.get("dimension")
            else:
                dimension = None  # 如果dimension为None，保持为None
        except Exception as e:
            logger.error(f"尺寸换算错误: {e}")
            dimension = None
        self.ws.cell(row=row_idx, column=12, value=weight)
        self.ws.cell(row=row_idx, column=13, value=dimension)

        # todo 卖家
        self.ws.cell(row=row_idx, column=14, value=item.get("seller_name", ""))

        # todo卖家地址
        businessAddress = ""
        try:
            seller_dto = item.get('sellerDto')
            print(item)
            if not seller_dto :
                seller_dto = item.get("seller_dto")
            if seller_dto:
                # 需要翻译
                business_address = seller_dto.get("businessAddress", "")
                if not business_address:
                    business_address = seller_dto.get("business_address", "")
                fanyi = BaiduTranslation().to_text(business_address.replace("<br/>", "\n"))
                if fanyi['trans_result']:
                    businessAddress = "".join([k['dst'] for k in fanyi['trans_result']])
        except Exception as e:
            logger.error(f'卖家地址解析失败！{asin} : {str(e)}')
        self.ws.cell(row=row_idx, column=15, value=businessAddress)

        # todo 补充信息
        item_info = ""
        try:
            relationKeyword = item.get("relationKeyword")
            if relationKeyword:
                item_info += f"全部流量词: {str(relationKeyword['total'])}\n"
                item_info += f"自然搜索词: {str(relationKeyword['nature'])}\n"
                item_info += f"广告流量词: {str(relationKeyword['ads'])}\n"
                item_info += f"搜索推荐词: {str(relationKeyword['recommend'])}\n"
        except Exception as e:
            logger.error(e)
        material = item.get("overviews", {})
        if material:
            materialJSON = json.loads(material)
            materialText = "; ".join([f"{k}: {v}" for k, v in materialJSON.items()])
            try:
                # 百度翻译
                fanyi = BaiduTranslation().to_text(materialText)
                if fanyi['trans_result']:
                    materialText = "".join([k['dst'] for k in fanyi['trans_result']])
                    item_info = f"材料信息: {materialText}\n" + item_info
            except Exception as e:
                logger.error(f'材料信息翻译失败：{e}')
        item_info = f"规格：{item.get("sku")}\n{item_info}"
        self.ws.cell(row=row_idx, column=16, value=item_info)

        # todo 同款信息
        try:
            similarList = json.loads(item.get("similarList"))
            if similarList:
                # 按评分排序
                sorted_data = sorted(similarList,
                                     key=lambda x: x.get('averageOverallRating', 0) or 0,
                                     reverse=True)
                fs = []
                for i in range(3):
                    fellow_info = sorted_data[i]
                    amazon_url = f"{item.get('itemWEB','')}/dp/{fellow_info.get('asin')}?th=1"
                    fs.append(f"{fellow_info.get('asin')}, 价格：{fellow_info.get('price', '')}, 评分{str(fellow_info.get('averageOverallRating'))} {fellow_info.get("totalReviewCount", '')}评论, 超链接: {amazon_url}")
                    # 同款图
                    item_image = fellow_info.get('imageUrl', '')
                    if item_image:
                        try:
                            img_data = self.download_image(item_image)
                            img = Image(img_data)
                            img.width = self.image_config['main_image']['width']
                            img.height = self.image_config['main_image']['height']
                            # 修复：直接设置anchor为单元格地址字符串
                            img.anchor = f"{chr(64 + 17 + i + 1)}{row_idx}"
                            self.ws.add_image(img)
                        except Exception as e:
                            logger.error(f"{asin}插入同款图失败: {e}")
                            self.ws.cell(row=row_idx, column=17 + i + 1, value='')
                similarText = '\n'.join(fs)
                self.ws.cell(row=row_idx, column=17, value=similarText)
            else:
                raise
        except Exception as e:
            logger.error(f'处理同款信息失败{asin}：{e}')
            self.ws.cell(row=row_idx, column=17, value="")

        # todo 预计配送费
        fba_txt = f"{currency}{item.get("fba", "")}"
        self.ws.cell(row=row_idx, column=21, value= fba_txt)

        # todo 五点描述
        Product_information = ""
        # 翻译
        try:
            description = item.get("description", "")
            if not description:
                raise Exception('description为空')
            PF = BaiduTranslation().to_text(description)
            if PF and PF.get('trans_result'):
                Product_information = "\n".join([k['dst'] for k in PF['trans_result']])
        except Exception as e:
            logger.error(f"获取描述失败 {asin}: {e}")
        self.ws.cell(row=row_idx, column=22, value=Product_information)

        # todo 阿里搜图
        ai_items = json.loads(item.get("aliexpress", []))
        if ai_items:
            # 图一
            image_1 = ai_items[0].get("imageUrl", "")
            try:
                img_data = self.download_image(image_1)
                img = Image(img_data)
                img.width = self.image_config['main_image']['width']
                img.height = self.image_config['main_image']['height']
                # 修复：直接设置anchor为单元格地址字符串
                img.anchor = f"{chr(64 + 23)}{row_idx}"
                self.ws.add_image(img)
            except Exception as e:
                logger.error(f"插入图1失败: {e}")

            # 图二
            image_2 = ai_items[1].get("imageUrl", "")
            try:
                img_data = self.download_image(image_2)
                img = Image(img_data)
                img.width = self.image_config['main_image']['width']
                img.height = self.image_config['main_image']['height']
                # 修复：直接设置anchor为单元格地址字符串
                img.anchor = f"{chr(64 + 24)}{row_idx}"
                self.ws.add_image(img)
            except Exception as e:
                logger.error(f"插入图2失败: {e}")

            # 图三
            image_3 = ai_items[2].get("imageUrl", "")
            try:
                img_data = self.download_image(image_3)
                img = Image(img_data)
                img.width = self.image_config['main_image']['width']
                img.height = self.image_config['main_image']['height']
                # 修复：直接设置anchor为单元格地址字符串
                img.anchor = f"{chr(64 + 25)}{row_idx}"
                self.ws.add_image(img)
            except Exception as e:
                logger.error(f"插入图3失败: {e}")

            # 采购价(元)
            price = "\n".join([k["price"] for k in ai_items[:5]])
            self.ws.cell(row=row_idx, column=26, value=price)

            # 采购商(1688)
            purchaser = "\n".join([k["companyName"] for k in ai_items[:5]])
            self.ws.cell(row=row_idx, column=27, value=purchaser)

            # 供应商1688链接
            link = "\n".join([k["link"] for k in ai_items[:5]])
            self.ws.cell(row=row_idx, column=28, value=link)
        else:
            self.ws.cell(row=row_idx, column=23, value="")
            self.ws.cell(row=row_idx, column=24, value="")
            self.ws.cell(row=row_idx, column=25, value="")
            self.ws.cell(row=row_idx, column=26, value="")
            self.ws.cell(row=row_idx, column=27, value="")
            self.ws.cell(row=row_idx, column=28, value="")


    @staticmethod
    def _format_bsr_ranking(bsr_list):
        """
            格式化BSR排名信息
            :param bsr_list: BSR列表
            :return: 格式化后的字符串
        """
        try:
            bsr_s = []
            if bsr_list:
                for bsr in bsr_list:
                    bsr_s.append(f"#{bsr['rank']} {bsr['label']};")
                return '\n'.join(bsr_s)
            return ""
        except Exception as e:
            logger.error(f"BSR_list: {e}")
            return ""

    def save(self):
        """保存工作簿"""
        if self.wb:
            # 设置所有单元格垂直居中
            for row in self.ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(vertical='center')
            self.wb.save(self.filename)
            logger.info(f"Excel文件已保存为: {self.filename}")
        else:
            logger.warning("没有工作簿需要保存")


    def close(self):
        """关闭工作簿"""
        if self.wb:
            self.wb.close()
            self.wb = None
            self.ws = None